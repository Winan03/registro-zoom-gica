import pandas as pd
import re
import unicodedata
from difflib import SequenceMatcher, get_close_matches
import requests
from datetime import datetime, date
import difflib
import os
import json
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl import load_workbook

# Constantes y configuraciones
titulos_academicos = ["mtr", "msc", "lic", "sr", "srta", "dra", "est", "prof", "doc", "aux"]
siglas_universidades = [
    "unprg", "unsa", "unmsm", "unt", "utp", "ucv", "upao", "ulima", "udep", "unfv",
    "ucsm", "upsjb", "uigv", "unp", "uni", "usmp", "unsaac", "uancv"
]
otros_prefijos = [
    "ix", "x", "piu", "chi", "tru", "sal", "invitado", "participante", "usuario", "alumno"
]

regex_prefijos = r'^(?:' + '|'.join(titulos_academicos + siglas_universidades + otros_prefijos) + r')[\s\-\_]+'
TITULOS_PERSONALES = {"ing", "dr", "dra", "lic", "mtr", "msc", "sr", "srta"}

# Variables globales que deben ser gestionadas desde el sistema principal
nombre_original_df = pd.DataFrame()
areas_disponibles = []
fechas_disponibles_var = []
fechas_seleccionadas_var = []

# Agregar estas variables globales al inicio del archivo
df_actual_filtrado = pd.DataFrame()  # Para almacenar el estado actual de los filtros
estado_filtros = {
    'area': 'TODOS',
    'fechas': [],
    'turno': 'TODOS',
    'busqueda': ''
}

# Cargar JSON de practicantes
url_json = "https://bucketreportezoom.s3.us-east-1.amazonaws.com/ultimo.json"
try:
    response = requests.get(url_json)
    response.raise_for_status()
    data_json = response.json()
    json_areas = {}
    for p in data_json:
        nombre_norm = unicodedata.normalize('NFD', p['nombre']).encode('ascii', 'ignore').decode('utf-8').lower().strip()
        partes = nombre_norm.split()
        if len(partes) < 3:
            continue
        json_areas[nombre_norm] = p['area']
        if len(partes) > 1:
            nombre_invertido = ' '.join(reversed(partes))
            json_areas[nombre_invertido] = p['area']
except Exception as e:
    print(f"Error al cargar JSON: {e}")
    data_json = []
    json_areas = {}

def normalizar_nombre(nombre):
    """Normaliza nombres eliminando prefijos institucionales y académicos sin dañar nombres válidos."""
    if not isinstance(nombre, str):
        return ""
    
    nombre = nombre.lower()
    nombre_original = nombre

    # Quitar paréntesis con contenido
    nombre = re.sub(r'\(.*?\)', '', nombre)

    # Eliminar una vez prefijos seguros (solo si están claramente al inicio con guión o espacio)
    nombre = re.sub(regex_prefijos, '', nombre)

    # Evitar errores como "ingrid", "luis", etc.
    nombre = re.sub(r'^[a-z]{2,6}[\-\_]', '', nombre)

    # Normalizar tildes y símbolos
    nombre = unicodedata.normalize('NFD', nombre)
    nombre = ''.join(c for c in nombre if unicodedata.category(c) != 'Mn')

    # Eliminar caracteres raros
    nombre = re.sub(r'[^a-z\s]', '', nombre)
    nombre = re.sub(r'\s+', ' ', nombre).strip()

    # Restaurar si quedó vacío
    if not nombre:
        return nombre_original.lower()

    # Reducir nombres muy largos a 3 componentes
    palabras = nombre.split()
    if len(palabras) > 4:
        palabras = [palabras[0]] + palabras[-2:]

    return ' '.join(palabras)

def agrupar_nombres_similares(df):
    """
    Agrupa nombres similares con una validación robusta y jerárquica,
    considerando iniciales y componentes faltantes.
    """
    if 'nombre' not in df.columns:
        return df

    df['nombre_normalizado'] = df['nombre'].apply(normalizar_nombre)
    nombres_unicos = df['nombre_normalizado'].unique().tolist()
    mapeo_final = {}
    usados = set()

    for nombre_base_norm in nombres_unicos:
        if nombre_base_norm in usados:
            continue

        palabras_base = nombre_base_norm.split()
        set_base = set(palabras_base)
        grupo = [nombre_base_norm]

        for otro_nombre_norm in nombres_unicos:
            if otro_nombre_norm == nombre_base_norm or otro_nombre_norm in usados:
                continue

            palabras_otro = otro_nombre_norm.split()
            set_otro = set(palabras_otro)

            # --- Validación Jerárquica ---

            # Nivel 1: Coincidencia Exacta (después de normalizar)
            if nombre_base_norm == otro_nombre_norm:
                grupo.append(otro_nombre_norm)
                usados.add(otro_nombre_norm)
                continue

            # Nivel 2: Mismos Componentes (sin importar el orden, hasta 4 palabras)
            if len(palabras_base) <= 4 and len(palabras_otro) <= 4 and set_base == set_otro:
                grupo.append(otro_nombre_norm)
                usados.add(otro_nombre_norm)
                continue

            # Nivel 3: Contención Mutua Significativa (uno contiene casi todas las palabras del otro)
            if (len(intersection := set_base.intersection(set_otro)) > 0 and
                    (len(intersection) / len(set_base) > 0.7 or len(intersection) / len(set_otro) > 0.7)):
                grupo.append(otro_nombre_norm)
                usados.add(otro_nombre_norm)
                continue

            # Nivel 4: Coincidencia Fuerte del Último Apellido y Alta Similitud del Resto
            if palabras_base and palabras_otro and palabras_base[-1] == palabras_otro[-1]:
                nombre_base_resto = " ".join(palabras_base[:-1])
                nombre_otro_resto = " ".join(palabras_otro[:-1])
                if SequenceMatcher(None, nombre_base_resto, nombre_otro_resto).ratio() > 0.7:
                    grupo.append(otro_nombre_norm)
                    usados.add(otro_nombre_norm)
                    continue

            # Nivel 4.5: Mismo primer nombre y coincidencia parcial en apellidos
            if len(palabras_base) >= 2 and len(palabras_otro) >= 2:
                if palabras_base[0] == palabras_otro[0]:
                    # Comparar apellidos (últimos dos elementos si existen)
                    apellidos_base = set(palabras_base[1:])
                    apellidos_otro = set(palabras_otro[1:])
                    if len(apellidos_base.intersection(apellidos_otro)) >= 1:
                        grupo.append(otro_nombre_norm)
                        usados.add(otro_nombre_norm)
                        continue

            # Nivel 4.8: Ignorar título si está en la primera palabra y comparar desde la segunda
            if palabras_base and palabras_otro:
                base_sin_titulo = palabras_base.copy()
                otro_sin_titulo = palabras_otro.copy()

                if base_sin_titulo[0] in TITULOS_PERSONALES:
                    base_sin_titulo = base_sin_titulo[1:]
                if otro_sin_titulo[0] in TITULOS_PERSONALES:
                    otro_sin_titulo = otro_sin_titulo[1:]

                if len(base_sin_titulo) >= 2 and len(otro_sin_titulo) >= 2:
                    set_base_sin = set(base_sin_titulo)
                    set_otro_sin = set(otro_sin_titulo)
                    inter = set_base_sin.intersection(set_otro_sin)
                    union = set_base_sin.union(set_otro_sin)
                    if len(inter) >= 2 or (len(inter) / len(union)) >= 0.6:
                        grupo.append(otro_nombre_norm)
                        usados.add(otro_nombre_norm)
                        continue

            # Nivel 5: Alta Similitud de Secuencia General (para pequeñas variaciones tipográficas o de orden)
            if SequenceMatcher(None, nombre_base_norm, otro_nombre_norm).ratio() > 0.8:
                grupo.append(otro_nombre_norm)
                usados.add(otro_nombre_norm)
                continue

            # --- Nueva Validación: Manejo de Iniciales y Componentes Faltantes ---
            if palabras_base and palabras_otro and palabras_base[-1] == palabras_otro[-1]:  # Misma Apellido
                nombres_base_sin_iniciales = " ".join([p for p in palabras_base if len(p) > 1])
                nombres_otro_sin_iniciales = " ".join([p for p in palabras_otro if len(p) > 1])
                if SequenceMatcher(None, nombres_base_sin_iniciales, nombres_otro_sin_iniciales).ratio() > 0.7:
                    grupo.append(otro_nombre_norm)
                    usados.add(otro_nombre_norm)
                    continue

            # --- Validaciones Negativas (para evitar agrupaciones incorrectas) ---
            # Validación para nombres con el mismo nombre y primer apellido pero diferente segundo apellido
            if len(palabras_base) >= 3 and len(palabras_otro) >= 3 and \
                    palabras_base[0] == palabras_otro[0] and palabras_base[1] == palabras_otro[1] and \
                    palabras_base[2] != palabras_otro[2]:
                continue

            # Validación para nombres con los mismos apellidos pero diferente primer nombre
            if len(palabras_base) >= 2 and len(palabras_otro) >= 2 and \
                    palabras_base[-2:] == palabras_otro[-2:] and palabras_base[0] != palabras_otro[0]:
                continue

        if grupo:
            nombre_representativo = max(grupo, key=lambda x: len(x))
            for item in grupo:
                mapeo_final[item] = nombre_representativo
                usados.add(item)

    df['nombre'] = df['nombre_normalizado'].map(mapeo_final).fillna(df['nombre_normalizado']).str.upper()
    return df

def obtener_nombre_completo_bd(df):
    if 'nombre' not in df.columns:
        return df
    mapa_nombres_completos = {}
    for p in data_json:
        nombre_norm = normalizar_nombre(p['nombre'])
        mapa_nombres_completos[nombre_norm] = p['nombre']
        partes = nombre_norm.split()
        if len(partes) > 1:
            mapa_nombres_completos[' '.join(reversed(partes))] = p['nombre']
    def encontrar_nombre_completo(nombre):
        nombre_norm = normalizar_nombre(nombre)
        if nombre_norm in mapa_nombres_completos:
            return mapa_nombres_completos[nombre_norm].upper()
        palabras = set(nombre_norm.split())
        for k, v in mapa_nombres_completos.items():
            if len(palabras.intersection(set(k.split()))) >= 2:
                return v.upper()
        matches = get_close_matches(nombre_norm, list(mapa_nombres_completos.keys()), n=1, cutoff=0.6)
        if matches:
            return mapa_nombres_completos[matches[0]].upper()
        return nombre.upper()
    df['nombre'] = df['nombre'].apply(encontrar_nombre_completo)
    return df

def buscar_area(nombre_practicante):
    """Busca el área de un practicante incluso si el nombre está parcial o desordenado."""
    nombre_norm = normalizar_nombre(nombre_practicante)
    
    # Asegurarse de que json_areas esté disponible (usando variable global)
    global json_areas
    
    # 1. Intentar coincidencia exacta
    if nombre_norm in json_areas:
        return json_areas[nombre_norm]
    
    # 2. Comparar con cada nombre en el json usando diferentes métodos
    for nombre_json, area in json_areas.items():
        nombre_json_norm = normalizar_nombre(nombre_json)
        
        # Comprobar si todas las palabras importantes están incluidas
        palabras_input = set(nombre_norm.split())
        palabras_json = set(nombre_json_norm.split())
        
        # Si hay una gran coincidencia en las palabras (más del 70%)
        palabras_comunes = palabras_input & palabras_json
        total_palabras = min(len(palabras_input), len(palabras_json))
        
        if total_palabras > 0 and len(palabras_comunes) / total_palabras >= 0.7:
            return area
        
        # Coincidencia por apellidos (suponiendo que están al final)
        if len(palabras_input) >= 2 and len(palabras_json) >= 2:
            apellidos_input = ' '.join(sorted(list(palabras_input))[-2:])
            apellidos_json = ' '.join(sorted(list(palabras_json))[-2:])

            if apellidos_input == apellidos_json:
                return area
        
        # Validación adicional: Primer nombre y primer apellido coinciden, pero segundo apellido es diferente
        nombres_input = list(palabras_input)
        nombres_json = list(palabras_json)
        
        if len(nombres_input) >= 2 and len(nombres_json) >= 2:
            primer_nombre_input, primer_apellido_input = nombres_input[0], nombres_input[1]
            primer_nombre_json, primer_apellido_json = nombres_json[0], nombres_json[1]
            
            if primer_nombre_input == primer_nombre_json and primer_apellido_input == primer_apellido_json:
                # Verifica el segundo apellido si existe
                if len(nombres_input) > 2 and len(nombres_json) > 2:
                    segundo_apellido_input = nombres_input[2]  # Usamos índice 2 porque es el tercer elemento
                    segundo_apellido_json = nombres_json[2]  # Similar aquí
                    
                    if segundo_apellido_input != segundo_apellido_json:
                        continue

            # Si los nombres coinciden pero el segundo apellido no, no agrupar
            if len(nombres_input) > 2 and len(nombres_json) > 2:
                if nombres_input[2] != nombres_json[2]:
                    continue

    # 3. Usar difflib para encontrar coincidencias cercanas
    matches = difflib.get_close_matches(
        nombre_norm,
        [normalizar_nombre(key) for key in json_areas.keys()],
        n=1,
        cutoff=0.6  # Reducir para mayor tolerancia
    )
    
    if matches:
        # Buscar el nombre original que coincide con el normalizado
        for nombre_json, area in json_areas.items():
            if normalizar_nombre(nombre_json) == matches[0]:
                return area
    
    return 'OTROS' 


def calcular_total_horas(grupo):
    """Calcula el total de horas trabajadas por un grupo de registros continuos."""
    grupo = grupo.sort_values('entrada')
    total = pd.Timedelta(0)
    i = 0

    while i < len(grupo):
        inicio = grupo.iloc[i]['entrada']
        fin = grupo.iloc[i]['salida']
        j = i + 1

        while j < len(grupo):
            siguiente_inicio = grupo.iloc[j]['entrada']

            # Tolerancia: 1 minuto máximo de diferencia entre fin y siguiente inicio
            if abs((siguiente_inicio - fin).total_seconds()) <= 60:
                fin = grupo.iloc[j]['salida']
                j += 1
            else:
                break

        total += (fin - inicio)
        i = j

    minutos_totales = round(total.total_seconds() / 60)  # 🔧 redondear
    horas_exactas = round(minutos_totales / 60, 2)        # 🔧 mantener redondeo exacto

    return f"{horas_exactas:.2f}", minutos_totales


def procesar_excel(file_paths):
    global nombre_original_df, areas_disponibles

    df_total = pd.DataFrame()

    for path in file_paths:
        if path.endswith(".csv"):
            df = pd.read_csv(path)
        else:
            df = pd.read_excel(path)

        required_columns = {'Nombre (nombre original)', 'Hora de entrada', 'Hora de salida'}
        if not required_columns.issubset(df.columns):
            continue

        df = df.rename(columns={
            'Nombre (nombre original)': 'nombre',
            'Hora de entrada': 'entrada',
            'Hora de salida': 'salida'
        }).copy()
        df = df[['nombre', 'entrada', 'salida']].dropna()
        df['nombre'] = df['nombre'].astype(str)
        df['entrada'] = pd.to_datetime(df['entrada'], dayfirst=True, errors='coerce')
        df['salida'] = pd.to_datetime(df['salida'], dayfirst=True, errors='coerce')

        df = df.dropna(subset=['entrada', 'salida'])
        df['fecha'] = df['entrada'].dt.date

        df_total = pd.concat([df_total, df], ignore_index=True)

    if df_total.empty:
        return pd.DataFrame()

    df_total['nombre_normalizado'] = df_total['nombre'].apply(normalizar_nombre)
    df_total = agrupar_nombres_similares(df_total)
    df_total = obtener_nombre_completo_bd(df_total)
    df_total['Área'] = df_total['nombre'].apply(buscar_area)

    nombre_original_df = df_total.copy()
    areas_disponibles = ['TODOS'] + sorted(df_total['Área'].unique().tolist())
    fechas = sorted(list(set(f.strftime('%d/%m/%Y') for f in df_total['fecha'])))

    fechas_disponibles_var.clear()
    fechas_disponibles_var.extend(fechas)
    fechas_seleccionadas_var.clear()

    ULTIMAS_RUTAS_CARGADAS.clear()
    ULTIMAS_RUTAS_CARGADAS.extend(file_paths)

    registrar_historial_carga_archivos(file_paths)  

    print(f"✅ Datos procesados: {len(df_total)} registros")

    df_total.to_pickle("datos_temporales.pkl")

    return generar_reporte(df_total)

def obtener_dataframe_vacio():
    """Retorna un DataFrame vacío para usar en Flask."""
    return pd.DataFrame()

def detectar_vacios(grupo):
    """
    Detecta vacíos entre registros consecutivos (>10 minutos) y retorna una lista de huecos
    """
    grupo_ordenado = grupo.sort_values(by='entrada')
    vacios = []
    for i in range(1, len(grupo_ordenado)):
        salida_anterior = grupo_ordenado.iloc[i - 1]['salida']
        entrada_actual = grupo_ordenado.iloc[i]['entrada']
        diferencia = entrada_actual - salida_anterior
        if diferencia.total_seconds() > 600:  # 10 minutos
            vacios.append({
                "salida": salida_anterior.strftime("%H:%M:%S"),
                "reingreso": entrada_actual.strftime("%H:%M:%S"),
                "duracion": str(diferencia).split('.')[0],  # Remover microsegundos
                "diferencia_segundos": diferencia.total_seconds()
            })
    return vacios

def calcular_horas_consideradas_vs_reales(grupo, vacios_detectados):
    """
    Calcula las horas consideradas (rango completo) vs horas reales (descontando vacíos)
    """
    if grupo.empty:
        return {}, {}
    
    # Ordenar por entrada
    grupo_ordenado = grupo.sort_values(by='entrada')
    
    # Hora de entrada más temprana y salida más tardía
    primera_entrada = grupo_ordenado.iloc[0]['entrada']
    ultima_salida = grupo_ordenado.iloc[-1]['salida']
    
    # Calcular tiempo total considerado
    tiempo_considerado = ultima_salida - primera_entrada
    
    # Calcular tiempo real (descontando vacíos)
    tiempo_vacios = sum([v['diferencia_segundos'] for v in vacios_detectados])
    tiempo_real_segundos = tiempo_considerado.total_seconds() - tiempo_vacios
    
    horas_consideradas = {
        'inicio': primera_entrada.strftime("%H:%M:%S"),
        'fin': ultima_salida.strftime("%H:%M:%S"),
        'total': str(tiempo_considerado).split(' ')[-1].split('.')[0]
    }
    
    horas_reales = {
        'total_segundos': tiempo_real_segundos,
        'total_formateado': str(pd.Timedelta(seconds=tiempo_real_segundos)).split(' ')[-1].split('.')[0]
    }
    
    return horas_consideradas, horas_reales

def generar_reporte(df):
    if df.empty:
        return pd.DataFrame()
    
    df = df.copy()
    df['fecha'] = pd.to_datetime(df['fecha'], errors='coerce').dt.date
    df.sort_values(by='fecha', inplace=True)
    reporte_final = []
    
    for fecha, grupo_fecha in df.groupby('fecha'):
        reporte_final.append({
            '#': '',
            'Nombre Practicante': f'📅 Fecha de Reporte: {fecha.strftime("%d/%m/%Y")}',
            'Turno Mañana': '', 'Turno Tarde': '',
            'Horas T.': '', 'Minutos Totales': '', 'Área': '',
            'Estado': '',  # Nueva columna
            'vacios_info': [],  # Info adicional para vacíos
            'fecha': fecha
        })
        
        grupo_fecha = grupo_fecha[~grupo_fecha['nombre'].str.startswith('📅 Fecha de Reporte')]
        
        for nombre, grupo in grupo_fecha.groupby('nombre'):
            # Detectar vacíos para este practicante
            vacios_detectados = detectar_vacios(grupo)
            
            turno_manana = {"entrada": None, "salida": None}
            turno_tarde = {"entrada": None, "salida": None}
            
            for _, row in grupo.iterrows():
                if row['entrada'].time() < datetime.strptime("14:00", "%H:%M").time():
                    turno_manana["entrada"] = min(turno_manana["entrada"], row['entrada']) if turno_manana["entrada"] else row['entrada']
                    turno_manana["salida"] = max(turno_manana["salida"], row['salida']) if turno_manana["salida"] else row['salida']
                else:
                    turno_tarde["entrada"] = min(turno_tarde["entrada"], row['entrada']) if turno_tarde["entrada"] else row['entrada']
                    turno_tarde["salida"] = max(turno_tarde["salida"], row['salida']) if turno_tarde["salida"] else row['salida']
            
            horas_t, minutos_total = calcular_total_horas(grupo)
            
            # Calcular horas consideradas vs horas reales
            horas_consideradas, horas_reales = calcular_horas_consideradas_vs_reales(grupo, vacios_detectados)
            
            def formatear_turno(turno):
                if turno["entrada"] and turno["salida"]:
                    return f"{turno['entrada'].strftime('%I:%M %p')} - {turno['salida'].strftime('%I:%M %p')}"
                return "NO INGRESO"
            
            # Determinar el estado basado en vacíos detectados
            estado_icono = "✅" if not vacios_detectados else "⚠️"
            
            reporte_final.append({
                '#': len([r for r in reporte_final if r['#'] != '']),
                'Nombre Practicante': nombre.upper(),
                'Turno Mañana': formatear_turno(turno_manana),
                'Turno Tarde': formatear_turno(turno_tarde),
                'Horas T.': horas_t,
                'Minutos Totales': minutos_total,
                'Área': grupo['Área'].iloc[0],
                'Estado': estado_icono,
                'vacios_info': {
                    'vacios': vacios_detectados,
                    'horas_consideradas': horas_consideradas,
                    'horas_reales': horas_reales,
                    'nombre': nombre.upper()
                },
                'fecha': fecha
            })
    
    return pd.DataFrame(reporte_final)


def filtrar_por_fechas_seleccionadas():
    if nombre_original_df.empty:
        return pd.DataFrame()

    fechas_dt = [datetime.strptime(f, "%d/%m/%Y").date() for f, v in fechas_seleccionadas_var if v]
    df_filtrado = nombre_original_df[nombre_original_df['fecha'].isin(fechas_dt)].copy()

    return generar_reporte(df_filtrado)


def aplicar_filtros(area, fechas, turno):
    """Aplica los filtros seleccionados y genera el reporte filtrado."""
    global nombre_original_df
    
    if nombre_original_df.empty:
        return pd.DataFrame()
    
    df_filtrado_raw = nombre_original_df.copy()
    
    # Aplicar filtro de área
    if area != "TODOS":
        df_filtrado_raw = df_filtrado_raw[df_filtrado_raw['Área'] == area]
    
    # Aplicar filtro de fechas
    if fechas:  # Si hay fechas seleccionadas
        try:
            # Convertir fechas del formato que viene del frontend
            fechas_dt = []
            for f in fechas:
                if isinstance(f, str):
                    # Intentar diferentes formatos de fecha
                    try:
                        fecha_dt = datetime.strptime(f, "%d/%m/%Y").date()
                    except ValueError:
                        try:
                            fecha_dt = datetime.strptime(f, "%Y-%m-%d").date()
                        except ValueError:
                            continue
                    fechas_dt.append(fecha_dt)
            
            if fechas_dt:
                df_filtrado_raw = df_filtrado_raw[df_filtrado_raw['fecha'].isin(fechas_dt)]
        except Exception as e:
            print(f"Error al procesar fechas: {e}")
    
    # Generar el reporte después de aplicar los filtros iniciales
    df_exportar = generar_reporte(df_filtrado_raw.copy())
    
    # Aplicar filtro de turno sobre el reporte generado
    if turno != "TODOS":
        turno_seleccionado = turno.upper()
        if turno_seleccionado == "MAÑANA":
            columna_turno = 'Turno Mañana'
        elif turno_seleccionado == "TARDE":
            columna_turno = 'Turno Tarde'
        else:
            return df_exportar  # Si el turno no es válido, devolver sin filtrar
        
        # Filtrar por el turno seleccionado
        if columna_turno in df_exportar.columns:
            df_exportar = df_exportar[df_exportar[columna_turno] != "NO INGRESO"]
    
    return df_exportar

def obtener_filtros_actuales():
    """Obtiene los filtros disponibles basados en los datos cargados."""
    global nombre_original_df
    
    if nombre_original_df.empty:
        return [], []
    
    # Obtener áreas únicas
    areas = ['TODOS'] + sorted(nombre_original_df['Área'].unique().tolist())
    
    # Obtener fechas únicas y formatearlas
    fechas_disponibles = []
    if 'fecha' in nombre_original_df.columns:
        fechas_unicas = sorted(nombre_original_df['fecha'].unique())
        for fecha in fechas_unicas:
            if pd.notna(fecha):
                if isinstance(fecha, str):
                    fechas_disponibles.append(fecha)
                else:
                    fechas_disponibles.append(fecha.strftime("%d/%m/%Y"))
    
    return fechas_disponibles, areas

def exportar_reporte_con_filtros_actuales():
    """
    Exporta el reporte con los filtros que están actualmente aplicados en la interfaz
    """
    global df_actual_filtrado, estado_filtros
    
    if df_actual_filtrado.empty:
        # Si no hay datos filtrados actuales, generar con los filtros guardados
        return aplicar_filtros(
            estado_filtros['area'], 
            estado_filtros['fechas'], 
            estado_filtros['turno']
        )
    
    return df_actual_filtrado.copy()

def exportar_reporte_completo():
    """Exporta el reporte completo sin filtros"""
    global nombre_original_df
    
    if nombre_original_df.empty:
        return pd.DataFrame()
    
    return generar_reporte(nombre_original_df.copy())

def exportar_a_archivo(df_exportar, ruta):
    """Exporta el DataFrame a un archivo Excel o CSV con formato."""
    try:
        if ruta.endswith(".csv"):
            df_exportar.to_csv(ruta, index=False)
        else:
            df_exportar.to_excel(ruta, index=False)
            
            # Aplicar formato al archivo Excel
            wb = load_workbook(ruta)
            ws = wb.active
            
            border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'),
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            fill = PatternFill(start_color="DCE6F1", fill_type="solid")
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if cell.row == 1:
                        cell.fill = fill
                        cell.font = Font(bold=True)
            
            # Ajustar ancho de columnas
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 3, 50)
            
            wb.save(ruta)
        
        return True
    except Exception as e:
        print(f"Error al exportar archivo: {e}")
        return False

def quitar_tildes_y_ñ(texto):
    texto_sin_tildes = ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    return texto_sin_tildes.replace('ñ', 'n').replace('Ñ', 'N')

def buscar_por_nombre(texto_busqueda, df_original):
    """
    Función de búsqueda mejorada con mejor manejo de casos edge y debugging
    """
    # Validaciones iniciales
    if not texto_busqueda or not texto_busqueda.strip():
        print("⚠️ Texto de búsqueda vacío")
        return df_original.copy() if not df_original.empty else pd.DataFrame()
    
    if df_original.empty:
        print("⚠️ DataFrame original está vacío")
        return pd.DataFrame()
    
    # Limpiar y preparar texto de búsqueda
    texto_busqueda = texto_busqueda.strip()
    texto_busqueda_limpio = quitar_tildes_y_ñ(texto_busqueda.lower())
    palabras_clave = [p for p in texto_busqueda_limpio.split() if len(p) > 1]  # Filtrar palabras muy cortas
    
    print(f"🔍 Búsqueda: '{texto_busqueda}' -> '{texto_busqueda_limpio}'")
    print(f"📊 Registros disponibles: {len(df_original)}")
    print(f"🔑 Palabras clave: {palabras_clave}")
    
    if not palabras_clave:
        print("⚠️ No hay palabras clave válidas")
        return df_original.copy()
    
    # Crear copia del DataFrame
    df_trabajo = df_original.copy()
    
    # Función de coincidencia mejorada
    def coincide_busqueda(nombre):
        if pd.isna(nombre):
            return False
        
        nombre_limpio = quitar_tildes_y_ñ(str(nombre).lower())
        
        # Método 1: Todas las palabras deben estar presentes
        coincidencias = 0
        for palabra in palabras_clave:
            if palabra in nombre_limpio:
                coincidencias += 1
        
        # Si todas las palabras coinciden, es una búsqueda perfecta
        if coincidencias == len(palabras_clave):
            return True
        
        # Si al menos el 70% de las palabras coinciden y hay más de una palabra
        if len(palabras_clave) > 1 and coincidencias >= (len(palabras_clave) * 0.7):
            return True
        
        # Para búsquedas de una sola palabra, ser más flexible
        if len(palabras_clave) == 1 and coincidencias > 0:
            return True
        
        return False
    
    # Aplicar filtro principal
    mask_principal = df_trabajo['nombre'].apply(coincide_busqueda)
    df_filtrado = df_trabajo[mask_principal].copy()
    
    print(f"✅ Resultados con filtro principal: {len(df_filtrado)}")
    
    # Si no hay resultados, intentar búsqueda más flexible
    if df_filtrado.empty:
        print("🔄 Aplicando búsqueda flexible...")
        
        def busqueda_flexible(nombre):
            if pd.isna(nombre):
                return False
            
            nombre_limpio = quitar_tildes_y_ñ(str(nombre).lower())
            
            # Buscar coincidencias parciales más amplias
            for palabra in palabras_clave:
                # Coincidencia parcial: la palabra está contenida o contiene parte del nombre
                if len(palabra) >= 3:  # Solo para palabras de al menos 3 caracteres
                    if palabra in nombre_limpio or any(palabra in parte for parte in nombre_limpio.split()):
                        return True
                    
                    # Búsqueda por similitud de secuencia
                    for parte_nombre in nombre_limpio.split():
                        if len(parte_nombre) >= 3:
                            similitud = SequenceMatcher(None, palabra, parte_nombre).ratio()
                            if similitud >= 0.8:  # 80% de similitud
                                return True
            
            return False
        
        mask_flexible = df_trabajo['nombre'].apply(busqueda_flexible)
        df_filtrado = df_trabajo[mask_flexible].copy()
        
        print(f"🎯 Resultados con búsqueda flexible: {len(df_filtrado)}")
    
    # Si aún no hay resultados, mostrar información de debug
    if df_filtrado.empty:
        print("❌ No se encontraron resultados")
        print("📝 Algunos nombres disponibles para referencia:")
        
        nombres_unicos = df_trabajo['nombre'].unique()
        nombres_muestra = nombres_unicos[:10] if len(nombres_unicos) > 10 else nombres_unicos
        
        for i, nombre in enumerate(nombres_muestra, 1):
            print(f"   {i}. {nombre}")
        
        if len(nombres_unicos) > 10:
            print(f"   ... y {len(nombres_unicos) - 10} nombres más")
    
    return df_filtrado

# Función principal que debes usar en tu ruta Flask
def buscar_y_generar_reporte_con_estado(texto_busqueda, df_original):
    global df_actual_filtrado
    
    actualizar_estado_filtros(busqueda=texto_busqueda)
    df_filtrado = buscar_por_nombre(texto_busqueda, df_original)
    df_reporte = generar_reporte(df_filtrado)
    df_actual_filtrado = df_reporte.copy()

    # ✅ Aquí aseguras que se registre correctamente
    registrar_historial_busqueda(texto_busqueda, df_filtrado)

    return df_reporte


def actualizar_estado_filtros(area=None, fechas=None, turno=None, busqueda=None):
    """Actualiza el estado global de los filtros aplicados"""
    global estado_filtros
    
    if area is not None:
        estado_filtros['area'] = area
    if fechas is not None:
        estado_filtros['fechas'] = fechas
    if turno is not None:
        estado_filtros['turno'] = turno
    if busqueda is not None:
        estado_filtros['busqueda'] = busqueda

def aplicar_filtros_y_guardar_estado(area, fechas, turno, busqueda=""):
    """
    Aplica los filtros y guarda el estado actual para exportación
    """
    global df_actual_filtrado, estado_filtros
    
    # Actualizar estado de filtros
    actualizar_estado_filtros(area, fechas, turno, busqueda)
    
    # Aplicar filtros
    df_resultado = aplicar_filtros(area, fechas, turno)
    
    # Si hay búsqueda, aplicarla también
    if busqueda and busqueda.strip():
        # Primero obtener los datos filtrados en formato raw
        df_raw_filtrado = obtener_datos_raw_filtrados(area, fechas)
        df_busqueda = buscar_por_nombre(busqueda, df_raw_filtrado)
        df_resultado = generar_reporte(df_busqueda)
        
        # Aplicar filtro de turno después de la búsqueda si es necesario
        if turno != "TODOS":
            df_resultado = aplicar_filtro_turno_a_reporte(df_resultado, turno)
    
    # Guardar el estado actual
    df_actual_filtrado = df_resultado.copy()

    registrar_historial_filtros_aplicados()
    
    return df_resultado

def aplicar_filtro_turno_a_reporte(df_reporte, turno):
    """Aplica filtro de turno a un reporte ya generado"""
    if turno == "TODOS":
        return df_reporte
    
    turno_seleccionado = turno.upper()
    if turno_seleccionado == "MAÑANA":
        columna_turno = 'Turno Mañana'
    elif turno_seleccionado == "TARDE":
        columna_turno = 'Turno Tarde'
    else:
        return df_reporte
    
    # Filtrar por el turno seleccionado
    if columna_turno in df_reporte.columns:
        return df_reporte[df_reporte[columna_turno] != "NO INGRESO"]
    
    return df_reporte

def obtener_estado_filtros_actual():
    """Devuelve el estado actual de los filtros"""
    return estado_filtros.copy()

def obtener_datos_raw_filtrados(area, fechas):
    """Obtiene los datos raw aplicando solo filtros de área y fecha"""
    global nombre_original_df
    
    if nombre_original_df.empty:
        return pd.DataFrame()
    
    df_filtrado_raw = nombre_original_df.copy()
    
    # Aplicar filtro de área
    if area != "TODOS":
        df_filtrado_raw = df_filtrado_raw[df_filtrado_raw['Área'] == area]
    
    # Aplicar filtro de fechas
    if fechas:
        try:
            fechas_dt = []
            for f in fechas:
                if isinstance(f, str):
                    try:
                        fecha_dt = datetime.strptime(f, "%d/%m/%Y").date()
                    except ValueError:
                        try:
                            fecha_dt = datetime.strptime(f, "%Y-%m-%d").date()
                        except ValueError:
                            continue
                    fechas_dt.append(fecha_dt)
            
            if fechas_dt:
                df_filtrado_raw = df_filtrado_raw[df_filtrado_raw['fecha'].isin(fechas_dt)]
        except Exception as e:
            print(f"Error al procesar fechas: {e}")
    
    return df_filtrado_raw



def limpiar_filtros_y_busqueda():
    """Limpia todos los filtros y la búsqueda"""
    global df_actual_filtrado, estado_filtros
    
    # Resetear estado de filtros
    estado_filtros = {
        'area': 'TODOS',
        'fechas': [],
        'turno': 'TODOS',
        'busqueda': ''
    }
    
    # Generar reporte completo
    df_completo = generar_reporte(nombre_original_df.copy()) if not nombre_original_df.empty else pd.DataFrame()
    df_actual_filtrado = df_completo.copy()
    
    return df_completo

def mostrar_tabla_agrupada_por_fecha(df):
    """
    Convierte un DataFrame en HTML agrupado por fecha con la nueva columna de Estado
    """
    if df.empty:
        return "<p>No hay datos para mostrar.</p>"

    html = """
    <table border='1' cellpadding='5' cellspacing='0'>
        <thead>
            <tr>
                <th>#</th>
                <th>Nombre Practicante</th>
                <th>Turno Mañana</th>
                <th>Turno Tarde</th>
                <th>Horas T.</th>
                <th>Minutos Totales</th>
                <th>Área</th>
                <th>Control</th>
            </tr>
        </thead>
        <tbody>
    """

    for index, row in df.iterrows():
        if str(row['#']).strip() == '':
            # Fila de fecha
            html += f"<tr style='background-color:#DCE6F1;font-weight:bold'><td colspan='8'>{row['Nombre Practicante']}</td></tr>"
        else:
            # Fila de datos
            estado_celda = ""
            if row['Estado'] == "⚠️":
                # Crear datos para el modal
                vacios_info = row.get('vacios_info', {})
                vacios_json = json.dumps(vacios_info) if vacios_info else '{}'
                estado_celda = f"<span class='warning-icon' onclick='mostrarModalVacios({vacios_json.replace(chr(39), chr(34))})' style='cursor:pointer; font-size:20px;' title='Detectados vacíos de tiempo'>⚠️</span>"
            else:
                estado_celda = f"<span style='font-size:20px; color:green;' title='Sin vacíos detectados'>✅</span>"
            
            html += "<tr>"
            html += f"<td>{row['#']}</td>"
            html += f"<td>{row['Nombre Practicante']}</td>"
            html += f"<td>{row['Turno Mañana']}</td>"
            html += f"<td>{row['Turno Tarde']}</td>"
            html += f"<td>{row['Horas T.']}</td>"
            html += f"<td>{row['Minutos Totales']}</td>"
            html += f"<td>{row['Área']}</td>"
            html += f"<td style='text-align:center;'>{estado_celda}</td>"
            html += "</tr>"

    html += "</tbody></table>"
    return html

def reiniciar_todo():
    """Elimina todos los datos cargados, filtros, búsquedas y variables globales"""
    global nombre_original_df, df_actual_filtrado
    global fechas_disponibles_var, fechas_seleccionadas_var, areas_disponibles
    global estado_filtros

    nombre_original_df = pd.DataFrame()
    df_actual_filtrado = pd.DataFrame()
    fechas_disponibles_var.clear()
    fechas_seleccionadas_var.clear()
    areas_disponibles.clear()
    estado_filtros = {
        'area': 'TODOS',
        'fechas': [],
        'turno': 'TODOS',
        'busqueda': ''
    }

    return pd.DataFrame()  # Devuelve un df vacío por si lo usas

# ==== Funciones de manejo de historial ====

HISTORIAL_REGISTROS = []
RUTA_HISTORIAL = "historial_registros.json"
ULTIMAS_RUTAS_CARGADAS = []

class RegistroHistorial:
    def __init__(self, descripcion, fecha, archivos, filtros_aplicados=None, df_estado=None):
        self.descripcion = descripcion
        self.fecha = fecha
        self.archivos = archivos
        self.filtros_aplicados = filtros_aplicados if filtros_aplicados is not None else {}
        self.df_estado = df_estado if df_estado is not None else []

    def to_dict(self):
        return {
            "descripcion": self.descripcion,
            "fecha": self.fecha,
            "archivos": self.archivos,
            "filtros_aplicados": self.filtros_aplicados,
            "df_estado": self.df_estado  # ✅ incluirlo al guardar
        }

    @classmethod
    def from_dict(cls, data):
        return cls(
            descripcion=data["descripcion"],
            fecha=data["fecha"],
            archivos=data["archivos"],
            filtros_aplicados=data.get("filtros_aplicados", {}),
            df_estado=data.get("df_estado", [])  # ✅ incluirlo al cargar
        )
    
def guardar_en_historial(descripcion, archivos=None, filtros=None, df_estado=None):
    global HISTORIAL_REGISTROS
    fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if archivos is None:
        archivos = []
    if filtros is None:
        filtros = {}

    # ✅ Convertir fechas en df_estado
    if df_estado is not None and isinstance(df_estado, pd.DataFrame):
        df_estado = df_estado.copy()
        for col in df_estado.columns:
            if df_estado[col].dtype == 'datetime64[ns]' or df_estado[col].dtype == 'object':
                df_estado[col] = df_estado[col].apply(
                    lambda x: x.strftime('%Y-%m-%d') if isinstance(x, (datetime, date)) else x
                )
        df_estado = df_estado.to_dict(orient='records')

    registro = RegistroHistorial(descripcion, fecha_actual, archivos, filtros, df_estado)
    HISTORIAL_REGISTROS.append(registro)
    guardar_historial_en_archivo()


def guardar_historial_en_archivo():
    try:
        with open(RUTA_HISTORIAL, 'w', encoding='utf-8') as f:
            json.dump([r.to_dict() for r in HISTORIAL_REGISTROS], f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error guardando historial: {e}")

def cargar_historial_desde_archivo():
    global HISTORIAL_REGISTROS
    try:
        if os.path.exists(RUTA_HISTORIAL):
            with open(RUTA_HISTORIAL, 'r', encoding='utf-8') as f:
                data = json.load(f)
                HISTORIAL_REGISTROS = [RegistroHistorial.from_dict(item) for item in data]
    except Exception as e:
        print(f"Error cargando historial: {e}")

def obtener_historial():
    try:
        if os.path.exists(RUTA_HISTORIAL):
            with open(RUTA_HISTORIAL, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"Error al obtener historial: {e}")
    return []


def registrar_historial_carga_archivos(file_paths):
    """Registra en el historial la carga de archivos"""
    if not file_paths:
        return
    
    descripcion = f"Carga de archivos: {', '.join([os.path.basename(p) for p in file_paths])}"
    guardar_en_historial(descripcion, archivos=file_paths)



def registrar_historial_filtros_aplicados():
    """Registra en el historial la aplicación de filtros"""
    filtros = obtener_estado_filtros_actual()
    descripcion = (
        f"Filtros aplicados: "
        f"Área={filtros['area'] or 'TODOS'}, "
        f"Fechas={filtros['fechas'] or 'NINGUNA'}, "
        f"Turno={filtros['turno'] or 'TODOS'}, "
        f"Búsqueda={filtros['busqueda'] or 'NINGUNA'}"
    )
    guardar_en_historial(descripcion, archivos=ULTIMAS_RUTAS_CARGADAS, filtros=filtros)

def restaurar_registro_desde_historial(registro_dict):
    global nombre_original_df, df_actual_filtrado

    archivos = registro_dict.get("archivos", [])
    filtros = registro_dict.get("filtros_aplicados", {})
    df_estado_raw = registro_dict.get("df_estado", [])

    df = pd.DataFrame()

    if archivos:
        df = procesar_excel(archivos)
        print("✅ Restauración ejecutada desde historial con archivos:", archivos)
        nombre_original_df = df.copy()

    elif df_estado_raw:
        df = pd.DataFrame(df_estado_raw)
        print(f"✅ Restauración ejecutada desde historial (df_estado), registros: {len(df)}")

        # Evita reprocesar, ya es un reporte
        df_actual_filtrado = df.copy()

    if 'nombre' not in df.columns and 'Nombre Practicante' in df.columns:
            # Ya es un reporte final, lo retornamos directamente
        return df

    # ⚠️ Solo aplica filtros si no hubo df_estado
    if filtros:
        area = filtros.get("area", "TODOS")
        fechas = filtros.get("fechas", [])
        turno = filtros.get("turno", "TODOS")
        busqueda = filtros.get("busqueda", "")
        df_restaurado = aplicar_filtros_y_guardar_estado(area, fechas, turno, busqueda)
        return df_restaurado

    return df


def registrar_historial_busqueda(texto, df_resultado):
    filtros = obtener_estado_filtros_actual()
    descripcion = f"Búsqueda realizada: '{texto}'"
    guardar_en_historial(descripcion, archivos=ULTIMAS_RUTAS_CARGADAS, filtros=filtros, df_estado=df_resultado)



